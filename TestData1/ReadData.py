import openpyxl

wk = openpyxl.load_workbook("E://TestData100.xlsx")


def get_no_rows(Sheetname):
    sh = wk[Sheetname]
    return sh.max_row


def fetch_cell_data(Sheetname, row, cell):
    sh = wk[Sheetname]
    cell = sh.cell(int(row), int(cell))
    return cell.value



